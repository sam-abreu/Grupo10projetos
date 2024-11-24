
from django.shortcuts import render, redirect
from .models import Doacao, Brinquedo, Inscricao, Profile
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font

def adminpage(request):
    return render(request, 'admin.html')

def home(request):
    return render(request, 'home.html')

def cadastro_user(request):

    mensagem = None
    tipo_mensagem = None

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = User.objects.filter(username=username).first()

        if user:
            mensagem = "Já existe um usuário com esse nome. Tente novamente."
            tipo_mensagem = "error"
            return render(request, 'cadastro_user.html', {'mensagem': mensagem, 'tipo_mensagem':tipo_mensagem})
        
        else:
            user = User.objects.create_user(username=username, password=password)
            user.save()
            mensagem = "Usuário criado com sucesso. Faça o login clicando no link abaixo."
            tipo_mensagem = "success"
            return render(request, 'cadastro_user.html', {'mensagem': mensagem, 'tipo_mensagem': tipo_mensagem})
        
    else:
        return render(request, 'cadastro_user.html')
    

def trabalhe_conosco(request):
    if request.method == 'POST':
        nome = request.POST['nome']
        email = request.POST['email']
        telefone = request.POST['telefone']
        mensagem = request.POST['mensagem']

        if nome and email and telefone and mensagem:

            inscricao = Inscricao(nome=nome, email=email, telefone=telefone, mensagem=mensagem)
            inscricao.save()

            condicao = "Sua inscrição foi enviado com sucesso!"
            tipo_condicao = "success"

        else:
            condicao = "Erro ao enviar inscrição. Preencha todos os campos e envie novamente"
            tipo_condicao = "error"

        return render(request, 'trabalhe_conosco.html', {'condicao': condicao, 'tipo_condicao':tipo_condicao})
    else:
        return render(request, 'trabalhe_conosco.html')
    
def visualizar_trabalheconosco(request):
    inscricoes = Inscricao.objects.all()
    return render(request, 'visualizar_trabalheconosco.html',{'inscricoes': inscricoes})

def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)

            if user.is_superuser:
                return redirect('adminpage')
            else:
                return redirect('home')
            
        else:
            mensagem = "Usuário ou senha inválido. Tente novamente."
            tipo_mensagem = "error"
            return render(request, 'login.html', {'mensagem': mensagem, 'tipo_mensagem': tipo_mensagem})

    else:
        return render(request, 'login.html')


# View para adicionar uma nova doação
@login_required
def add_doacao(request):
    if request.method == "POST":
        nome_doador = request.POST.get("nome_doador")
        email_doador = request.POST.get("email_doador")
        telefone_doador = request.POST.get("telefone_doador")
        tipo_doacao = request.POST.get("tipo_doacao")
        descricao = request.POST.get("descricao")

        # Criação de uma nova doação no banco de dados
        doacao = Doacao(
            nome_doador=nome_doador,
            email_doador=email_doador,
            telefone_doador=telefone_doador,
            tipo_doacao=tipo_doacao,
            descricao=descricao
        )
        doacao.save()

        profile = request.user.profile
        profile.pontos += 5
        profile.save()
        
        # Redireciona para uma página de confirmação ou para outra ação após salvar
        return redirect('home')

    return render(request, 'doar.html')

def residuos(request):
    brinquedos = Brinquedo.objects.all()
    return render(request, 'residuos.html', {'brinquedos': brinquedos})

# View para gerar o relatório em Excel
def gerar_relatorio_doacoes(request):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Relatório de Doações"

    headers = ["Nome do Doador", "Email do Doador", "Telefone do Doador", "Tipo de Doação", "Descrição", "Data da Doação"]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    doacoes = Doacao.objects.all()
    for row_num, doacao in enumerate(doacoes, 2):
        worksheet.cell(row=row_num, column=1, value=doacao.nome_doador)
        worksheet.cell(row=row_num, column=2, value=doacao.email_doador)
        worksheet.cell(row=row_num, column=3, value=doacao.telefone_doador)
        worksheet.cell(row=row_num, column=4, value=doacao.tipo_doacao)
        worksheet.cell(row=row_num, column=5, value=doacao.descricao)
        worksheet.cell(row=row_num, column=6, value=doacao.data_doacao.strftime('%d/%m/%Y %H:%M'))

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="relatorio_doacoes.xlsx"'
    workbook.save(response)
    return response

def visualizar_doacao(request):
    doacoes = Doacao.objects.all()
    return render(request, 'visualizar_doacao.html', {'doacoes': doacoes})


def add_brinquedo(request):
    if request.method == 'POST':
        nome = request.POST.get('nome')
        preco = request.POST.get('preco')
        faixaetaria = request.POST.get('faixaetaria')
        material = request.POST.get('material')

        # Cria um novo brinquedo e salva no banco de dados
        novo_brinquedo = Brinquedo(
            nome=nome,
            preco=preco,
            faixaetaria=faixaetaria,
            material=material
        )
        novo_brinquedo.save()

        # Redireciona para a página principal ou onde você desejar
        return redirect('home')  # Certifique-se de que 'home' é o nome correto da URL

    return render(request, 'add_brinquedo.html')  # Renderiza um template para adicionar brinquedos

@login_required
def comprar_brinquedo(request):
    brinquedos = Brinquedo.objects.all()  # Puxa todos os brinquedos do banco de dados
    profile = request.user.profile

    for brinquedo in brinquedos:
        brinquedo.preco_com_desconto = float(brinquedo.preco) - (profile.pontos / 100 * float(brinquedo.preco))

    context = {'brinquedos': brinquedos,
               'profile': profile
            }

    return render(request, 'comprar_brinquedo.html', context) 


def sugestao_brinquedo(request):
    idade = None
    sexo = ''
    sugestao = ''
    
    if request.method == 'POST':
        idade = float(request.POST.get('idade'))
        sexo = request.POST.get('sexo', '')

        if idade >= 0 and idade <= 2 and sexo == 'Masculino':
            sugestao = 'Boneco de herói'

        elif idade > 2 and idade <= 6 and sexo == 'Masculino':
            sugestao = 'Carro de brinquedo'
        
        elif idade > 6 and idade <= 10 and sexo == 'Masculino':
            sugestao = 'Brinquedo de construção'
        
        if idade >= 0 and idade <= 2 and sexo == 'Feminino':
            sugestao = 'Boneca'

        elif idade > 2 and idade <= 6 and sexo == 'Feminino':
            sugestao = 'Palácio de princesas'
        
        elif idade > 6 and idade <= 10 and sexo == 'Feminino':
            sugestao = 'Brinquedo de mercadinho'
        
        elif idade > 10 and sexo == 'Feminino':
            sugestao = 'Excedeu a idade limite'

        elif idade > 10 and sexo == 'Masculino':
            sugestao = 'Excedeu a idade limite'


    return render(request, 'sugestao_brinquedo.html', {
        'sexo': sexo,
        'idade': idade,
        'sugestao' : sugestao
    })